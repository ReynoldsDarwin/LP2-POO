import customtkinter as ctk
from tkinter import messagebox
from datetime import datetime
import random
import os
from openpyxl import Workbook, load_workbook

# ================== CLASES BASE ==================
class MetodoPago:
    def pago(self, cantidad, usuario):
        pass

class TarjetaDeCredito(MetodoPago):
    def __init__(self):
        self.nombre = "💳 Tarjeta de Crédito"
    def pago(self, cantidad, usuario):
        return f"{usuario} pagó ${cantidad:.2f} con {self.nombre}"

class PayPal(MetodoPago):
    def __init__(self):
        self.nombre = "💠 PayPal"
    def pago(self, cantidad, usuario):
        return f"{usuario} pagó ${cantidad:.2f} mediante {self.nombre}"

class PagoEfectivo(MetodoPago):
    def __init__(self):
        self.nombre = "💵 Efectivo"
    def pago(self, cantidad, usuario):
        return f"{usuario} pagó ${cantidad:.2f} en {self.nombre}"

class Yape(MetodoPago):
    def __init__(self):
        self.nombre = "📱 Yape"
    def pago(self, cantidad, usuario):
        return f"{usuario} pagó ${cantidad:.2f} usando {self.nombre}"


# ================== FUNCIONES GENERALES ==================
def generar_numero_operacion():
    return str(random.randint(100000, 999999))


def guardar_en_excel(pago):
    archivo = "historial_pagos.xlsx"
    if os.path.exists(archivo):
        wb = load_workbook(archivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Usuario", "DNI/RUC", "Monto", "Método", "Fecha y hora", "N° Operación"])

    ws.append([
        pago["usuario"],
        pago["dni"],
        pago["monto"],
        pago["metodo"],
        pago["fecha"],
        pago["operacion"]
    ])
    wb.save(archivo)


# ================== INTERFAZ PRINCIPAL ==================
def abrir_sistema():
    ventana_login.destroy()

    ventana = ctk.CTk()
    ventana.title("🏦 Banco Global Internacional - Sistema de Pagos")
    ventana.geometry("850x750")

    ctk.CTkLabel(ventana, text="🏦 BANCO GLOBAL INTERNACIONAL", font=("Segoe UI", 26, "bold")).pack(pady=10)
    ctk.CTkLabel(ventana, text="💳 Plataforma de Pagos Corporativos", font=("Segoe UI", 15, "italic")).pack(pady=5)

    main_frame = ctk.CTkScrollableFrame(ventana, corner_radius=15)
    main_frame.pack(fill="both", expand=True, padx=20, pady=15)

    # --- CAMPOS DE DATOS ---
    frame_datos = ctk.CTkFrame(main_frame, corner_radius=15)
    frame_datos.pack(pady=15, padx=15, fill="x")

    ctk.CTkLabel(frame_datos, text="👤 Nombre completo:", font=("Segoe UI", 14)).grid(row=0, column=0, padx=10, pady=10, sticky="e")
    entry_usuario = ctk.CTkEntry(frame_datos, placeholder_text="Ej: Juan Pérez", width=220)
    entry_usuario.grid(row=0, column=1, padx=10, pady=10)

    ctk.CTkLabel(frame_datos, text="🧾 DNI / RUC:", font=("Segoe UI", 14)).grid(row=1, column=0, padx=10, pady=10, sticky="e")
    entry_dni = ctk.CTkEntry(frame_datos, placeholder_text="Ej: 12345678 o 20654321098", width=220)
    entry_dni.grid(row=1, column=1, padx=10, pady=10)

    ctk.CTkLabel(frame_datos, text="💰 Monto ($):", font=("Segoe UI", 14)).grid(row=2, column=0, padx=10, pady=10, sticky="e")
    entry_monto = ctk.CTkEntry(frame_datos, placeholder_text="Ej: 100.50", width=220)
    entry_monto.grid(row=2, column=1, padx=10, pady=10)

    ctk.CTkLabel(frame_datos, text="💳 Método de pago:", font=("Segoe UI", 14)).grid(row=3, column=0, padx=10, pady=10, sticky="e")
    metodo_var = ctk.StringVar(value="Tarjeta de Crédito")
    menu_metodos = ctk.CTkOptionMenu(frame_datos, variable=metodo_var, values=["Tarjeta de Crédito", "PayPal", "Efectivo", "Yape"], width=220)
    menu_metodos.grid(row=3, column=1, padx=10, pady=10)

    # --- RECTÁNGULO DE RESULTADO MINIMALISTA ---
    canvas = ctk.CTkCanvas(main_frame, width=600, height=120, bg="#101010", highlightthickness=0)
    canvas.pack(pady=20)

    label_resultado = ctk.CTkLabel(main_frame, text="Resultado: ---", font=("Consolas", 16, "bold"), text_color="#CCCCCC")
    label_resultado.pack(pady=5)

    # --- ÚLTIMO COMPROBANTE ---
    frame_comprobante = ctk.CTkFrame(main_frame, corner_radius=10, fg_color="white")
    frame_comprobante.pack(pady=10, padx=20, fill="x")

    label_comprobante = ctk.CTkLabel(frame_comprobante, text="📜 Aquí aparecerá el último comprobante generado", 
                                     font=("Segoe UI", 13, "italic"), text_color="#333")
    label_comprobante.pack(pady=10)

    # --- HISTORIAL DEL DÍA ---
    frame_historial = ctk.CTkFrame(main_frame, corner_radius=15)
    frame_historial.pack(pady=20, padx=20, fill="x")

    ctk.CTkLabel(frame_historial, text="📊 HISTORIAL DEL DÍA", font=("Segoe UI", 17, "bold")).pack(pady=8)
    label_historial = ctk.CTkLabel(frame_historial, text="Aún no se han registrado operaciones.", 
                                   font=("Segoe UI", 13), text_color="#BBBBBB", justify="left")
    label_historial.pack(pady=8)

    historial_pagos = []  # lista local

    # --- FUNCIONES DEL SISTEMA ---
    def realizar_pago():
        usuario = entry_usuario.get().strip()
        dni = entry_dni.get().strip()
        monto = entry_monto.get().strip()

        if not usuario or not dni or not monto:
            messagebox.showwarning("Campos vacíos", "Por favor completa todos los campos antes de continuar.")
            return

        try:
            cantidad = float(monto)
            if cantidad <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Error", "El monto debe ser un número positivo.")
            return

        metodo = metodo_var.get()
        numero_operacion = generar_numero_operacion()
        fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if metodo == "Tarjeta de Crédito":
            metodo_pago = TarjetaDeCredito()
            color = "#2E86C1"
        elif metodo == "PayPal":
            metodo_pago = PayPal()
            color = "#16A085"
        elif metodo == "Efectivo":
            metodo_pago = PagoEfectivo()
            color = "#27AE60"
        else:
            metodo_pago = Yape()
            color = "#8E44AD"

        resultado = metodo_pago.pago(cantidad, usuario)
        label_resultado.configure(text=resultado)

        # Dibujo minimalista del pago
        canvas.delete("all")
        canvas.create_rectangle(180, 30, 420, 90, fill=color, outline="", width=0)
        canvas.create_text(300, 50, text=metodo_pago.nombre, fill="white", font=("Segoe UI", 13, "bold"))
        canvas.create_text(300, 75, text=f"${cantidad:.2f}", fill="white", font=("Consolas", 12))

        # Actualizar comprobante
        label_comprobante.configure(
            text=f"🏦 BANCO GLOBAL INTERNACIONAL\n"
                 f"👤 {usuario} | 🪪 {dni}\n"
                 f"💳 {metodo_pago.nombre} | 💵 ${cantidad:.2f}\n"
                 f"🕒 {fecha_actual}\n"
                 f"📄 N° Operación: {numero_operacion}",
            text_color="black", font=("Segoe UI", 13)
        )

        # Guardar en memoria y en Excel
        pago = {
            "usuario": usuario,
            "dni": dni,
            "monto": cantidad,
            "metodo": metodo_pago.nombre,
            "fecha": fecha_actual,
            "operacion": numero_operacion
        }
        historial_pagos.append(pago)
        guardar_en_excel(pago)

        # Actualizar historial del día
        total_operaciones = len(historial_pagos)
        total_monto = sum(p["monto"] for p in historial_pagos)
        metodos = {p["metodo"] for p in historial_pagos}

        resumen = (
            f"🧾 Operaciones realizadas: {total_operaciones}\n"
            f"💵 Monto total movido: ${total_monto:.2f}\n"
            f"💳 Métodos usados: {', '.join(metodos)}"
        )
        label_historial.configure(text=resumen)

        entry_monto.delete(0, "end")

    # --- BOTÓN DE PAGO ---
    ctk.CTkButton(main_frame, text="💳 Pagar", fg_color="#007ACC", hover_color="#005A8A",
                  width=160, height=40, command=realizar_pago).pack(pady=10)

    ctk.CTkLabel(main_frame, text="© 2025 Banco Global Internacional • Sistema Seguro SSL 🔐",
                 font=("Segoe UI", 11, "italic")).pack(pady=10)

    ventana.mainloop()


# ================== LOGIN ==================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

ventana_login = ctk.CTk()
ventana_login.title("🔐 Acceso al Sistema - Banco Global Internacional")
ventana_login.geometry("400x350")

ctk.CTkLabel(ventana_login, text="🏦 Acceso Seguro", font=("Segoe UI", 24, "bold")).pack(pady=25)

frame_login = ctk.CTkFrame(ventana_login, corner_radius=15)
frame_login.pack(pady=15, padx=30, fill="x")

ctk.CTkLabel(frame_login, text="👤 Usuario:", font=("Segoe UI", 14)).pack(pady=10)
entry_user = ctk.CTkEntry(frame_login, placeholder_text="admin", width=250)
entry_user.pack(pady=5)

ctk.CTkLabel(frame_login, text="🔑 Contraseña:", font=("Segoe UI", 14)).pack(pady=10)
entry_pass = ctk.CTkEntry(frame_login, placeholder_text="123456", show="•", width=250)
entry_pass.pack(pady=5)

def validar_login():
    if entry_user.get().strip() == "admin" and entry_pass.get().strip() == "123456":
        abrir_sistema()
    else:
        messagebox.showerror("Acceso denegado", "Usuario o contraseña incorrectos.")

ctk.CTkButton(frame_login, text="Ingresar 🔓", width=200, fg_color="#007ACC", hover_color="#005A8A",
              command=validar_login).pack(pady=20)
ctk.CTkLabel(ventana_login, text="© 2025 Banco Global Internacional - Acceso Autorizado",
             font=("Segoe UI", 10, "italic")).pack(pady=10)

ventana_login.mainloop()
